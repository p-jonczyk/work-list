import openpyxl as xl
from openpyxl.comments import Comment
import calendar
import sys
import os
import constants as const
import files_handeling
from datetime import time


def hour_representation_formula(hours, turn_int=True):
    """Creats formula to change number into number to be converted into HH:MM

    Parameters:

    work_hours (int): work hours of worker """

    my_time = int(hours*3600)
    hour_rep = time(my_time//3600, (my_time % 3600)//60).strftime("%H:%M")
    return hour_rep


def styling_rows(base_ws, day, current_row_number):
    """Styles rows based on day of week

    Parameters:

    base_ws : activated loaded base file

    day (int): day of week

    current_row_number (int): current row to be stylied """

    current_row_list = f"{current_row_number}:{current_row_number}"
    if day == 5 or day == 6:
        for cells in base_ws[current_row_list]:
            cells.style = "days_off"
    else:
        for cells in base_ws[current_row_list]:
            cells.style = "base"


def create_month_table(base_ws, year, month):
    """Creats table with rows number same as days in month
    and insert it into base file

    Parameter:

    base_ws : activated loaded base file

    year (int): current year

    month (int): order of current month """

    # insert proper quantity of rows based on days of current month and days off (color)
    cal = calendar.Calendar()
    for counter, day in cal.itermonthdays2(year, month):
        if counter != 0:
            # use 10 row + counter becuase we want to start with 11 row - counter starts from 1
            current_row_number = 10 + counter
            base_ws.insert_rows(current_row_number)
            # number each day of month from 1 to number of days
            current_cell = f'A{current_row_number}'
            base_ws[current_cell] = counter
            # marking  Saturdays and Sundays
            # first week day is 0 so 5th and 6th are Saturday and Sunday respectively
            styling_rows(base_ws, day, current_row_number)


def handle_nonhour_input(base_ws, value, work_start_cell, pause_cell, work_end_cell, work_hours_cell, comment_day_num_cell):
    """Handles different type of input in source file then hour(int)

    Parameters:

    base_ws: activated loaded base file

    value (str): nonhour input from source file

    work_start_cell (str): start hour cell reference

    pause_cell (str): pasue hour cell reference

    work_end_cell (str): end hour cell reference

    work_hours_cell (str): sum of work cell reference

    comment_day_num_cell (str): day number cell reference (for comment)"""

    # capitalizing proper legend input
    value_cap = value.upper()
    if value_cap in const.defined_cell_occurence:
        base_ws[work_start_cell] = value_cap
        base_ws[pause_cell] = value_cap
        base_ws[work_end_cell] = value_cap
        base_ws[work_hours_cell] = value_cap
    else:
        base_ws[comment_day_num_cell].comment = Comment(
            f"{value}", "")
        base_ws[work_start_cell] = ""
        base_ws[pause_cell] = ""
        base_ws[work_end_cell] = ""
        base_ws[work_hours_cell] = ""


def fill_rows(base_ws, row, max_col, work_hour_start_days, pause):
    """Fills each row with data obtained from source file

    Parameters:

    base_ws: activated loaded base file

    row (list): listo of values from current row

    max_col (int): number of columns in row (length of row list)

    work_hour_start_days (dict): dictionary of pairs {day_range_selection : work_start}

    pause (int/float): length of workers' break"""

    for j, value in enumerate(row[3:max_col]):
        work_start_cell = f'B{11+j}'
        pause_cell = f'C{11+j}'
        work_end_cell = f'D{11+j}'
        work_hours_cell = f'E{11+j}'
        # cell for comments
        comment_day_num_cell = f'A{11+j}'
        # takse proper work_start
        for day in work_hour_start_days.keys():
            if j < day:
                work_start = work_hour_start_days.get(day)
                break
        if value is not None:
            try:
                base_ws[work_hours_cell] = value
                base_ws[pause_cell] = pause
                value = int(value)
                base_ws[work_start_cell] = hour_representation_formula(
                    work_start)
                base_ws[work_end_cell] = hour_representation_formula(
                    work_start + pause + value)
            except ValueError:
                handle_nonhour_input(base_ws, value, work_start_cell, pause_cell,
                                     work_end_cell, work_hours_cell,
                                     comment_day_num_cell)


def fill_sign_date(base_ws, year, month, days_in_month):
    """Fills date of sign field in lower part of table

    Parameters:

    base_ws: activated loaded base file

    year (int): current year

    month (int): order of current month

    days_in_month (int): number of days in month """

    sign_cell = f'B{13+days_in_month}'
    if month < 10:
        base_ws[sign_cell] = (f'{days_in_month}.0{month}.{year}')
    else:
        base_ws[sign_cell] = (f'{days_in_month}.{month}.{year}')


def fill_worksheet(source_ws, base_ws, month_name, load_base, month, year, work_hour_start_days, pause, construction_site):
    """Filling worksheet with data from source file

    Parameters:

    source_ws: loaded source worksheet

    base_ws: activated loaded base file

    month_name (str): name of current month

    load_base: loaded base file (before activation)

    month (int): order of current month

    year (int): current year

    work_hour_start_days (dict): dictionary of pairs {day_range_selection : work_start}

    pause (int/float): length of workers' break

    construction_site (str): construction site name"""

    # get max row/col
    max_row = source_ws.max_row + 1
    max_col = source_ws.max_column - 1
    # getting number of days in current month
    days_in_month = calendar.monthrange(year, month)[1]

    # fill table with source file's data
    for row in source_ws.iter_rows(6, max_row, 1, max_col, values_only=True):

        create_month_table(base_ws, year, month)
        worker_number = row[0]
        worker_name = row[1]
        firm_name = row[2]
        # fills top information of file
        base_ws['G7'] = month_name
        base_ws['D7'] = worker_number
        base_ws['D5'] = worker_name
        base_ws['D3'] = firm_name
        # row filler
        fill_rows(base_ws, row, max_col, work_hour_start_days, pause)
        # fills date for sign place - last day of month
        fill_sign_date(base_ws, year, month, days_in_month)
        # setting sum formula for workhours of month
        base_ws[f'E{11+days_in_month}'] = (f"=SUM(E11:E{10+days_in_month})")
        # handles max_row reached
        if worker_name is not None:
            # saving data into new file with "Construction Site Name Surname" filename
            load_base.save(
                f".\{construction_site} employees worklists\{firm_name} {worker_name}.xlsx")
            # clearing added rows from base tamplate file
            base_ws.delete_rows(11, days_in_month)  # new
